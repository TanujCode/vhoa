import subprocess
import sys
try:
    import openpyxl
    import reportlab
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl==3.1.2", "reportlab==4.1.0"])

import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_rental_db
from app.routers.rental.dependencies import require_rental_role
from app.models.rental.rental_user import RentalUser
from app.models.rental.property import Property
from app.models.rental.unit import Unit
from app.models.rental.lease import Lease
from app.models.rental.rental_ledger import RentalLedger
from app.models.rental.rental_maintenance import RentalMaintenanceRequest
from app.models.rental.rental_vendor import RentalVendor

router = APIRouter(prefix="/rental", tags=["Rental - Reports"])


@router.get("/reports/stats")
def get_rental_report_stats(
    property_id: str = Query("all"),
    db: Session = Depends(get_rental_db),
    current_user: RentalUser = Depends(require_rental_role("super_admin", "landlord")),
):
    role_name = (current_user.role.role_name if current_user.role else "").lower()
    is_super = role_name == "super_admin"

    prop_query = db.query(Property).filter(Property.active_status == True)
    if not is_super:
        prop_query = prop_query.filter(Property.landlord_id == current_user.user_id)
    
    if property_id != "all":
        try:
            p_id = int(property_id)
            prop_query = prop_query.filter(Property.property_id == p_id)
        except ValueError:
            pass
            
    properties = prop_query.all()
    property_ids = [p.property_id for p in properties]
    
    if not property_ids:
        return {
            "leases": {"total": 0, "active": 0, "pending_signature": 0, "by_status": {}},
            "maintenance": {"total": 0, "by_status": {}, "total_cost": 0.0},
            "ledgers": {"total_count": 0, "total_collected": 0.0, "total_overdue": 0.0, "by_status": {}},
            "properties_count": 0,
            "units_count": 0,
            "tenants_count": 0
        }
        
    properties_count = len(property_ids)
    
    units_count = db.query(Unit).filter(Unit.property_id.in_(property_ids), Unit.active_status == True).count()
    
    tenants_count = db.query(Lease).filter(
        Lease.unit_id.in_(
            db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
        ),
        Lease.status == "ACTIVE"
    ).count()
    
    # Leases Stats
    leases_q = db.query(Lease).filter(
        Lease.unit_id.in_(
            db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
        )
    )
    total_leases = leases_q.count()
    active_leases = leases_q.filter(Lease.status == "ACTIVE").count()
    pending_sig_leases = leases_q.filter(Lease.status == "PENDING_SIGNATURE").count()
    
    lease_status_counts = (
        db.query(Lease.status, func.count(Lease.lease_id))
        .filter(Lease.unit_id.in_(
            db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
        ))
        .group_by(Lease.status)
        .all()
    )
    lease_status_summary = {status: count for status, count in lease_status_counts}
    
    # Maintenance Stats
    maint_q = db.query(RentalMaintenanceRequest).filter(
        RentalMaintenanceRequest.lease_id.in_(
            db.query(Lease.lease_id).filter(
                Lease.unit_id.in_(
                    db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                )
            )
        )
    )
    total_maint = maint_q.count()
    maint_status_counts = (
        db.query(RentalMaintenanceRequest.status, func.count(RentalMaintenanceRequest.request_id))
        .filter(RentalMaintenanceRequest.lease_id.in_(
            db.query(Lease.lease_id).filter(
                Lease.unit_id.in_(
                    db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                )
            )
        ))
        .group_by(RentalMaintenanceRequest.status)
        .all()
    )
    maint_status_summary = {status: count for status, count in maint_status_counts}
    
    total_maint_cost = maint_q.filter(RentalMaintenanceRequest.status == "COMPLETED").with_entities(func.sum(RentalMaintenanceRequest.estimated_cost)).scalar() or 0.0
    
    # Ledgers Stats
    ledgers_q = db.query(RentalLedger).filter(
        RentalLedger.lease_id.in_(
            db.query(Lease.lease_id).filter(
                Lease.unit_id.in_(
                    db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                )
            )
        )
    )
    total_ledgers_count = ledgers_q.count()
    total_collected = ledgers_q.filter(RentalLedger.status == "PAID").with_entities(func.sum(RentalLedger.amount)).scalar() or 0.0
    total_overdue = ledgers_q.filter(RentalLedger.status == "OVERDUE").with_entities(func.sum(RentalLedger.amount)).scalar() or 0.0
    
    ledger_status_counts = (
        db.query(RentalLedger.status, func.count(RentalLedger.invoice_id))
        .filter(RentalLedger.lease_id.in_(
            db.query(Lease.lease_id).filter(
                Lease.unit_id.in_(
                    db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                )
            )
        ))
        .group_by(RentalLedger.status)
        .all()
    )
    ledger_status_summary = {status: count for status, count in ledger_status_counts}
    
    return {
        "leases": {
            "total": total_leases,
            "active": active_leases,
            "pending_signature": pending_sig_leases,
            "by_status": lease_status_summary
        },
        "maintenance": {
            "total": total_maint,
            "by_status": maint_status_summary,
            "total_cost": total_maint_cost
        },
        "ledgers": {
            "total_count": total_ledgers_count,
            "total_collected": total_collected,
            "total_overdue": total_overdue,
            "by_status": ledger_status_summary
        },
        "properties_count": properties_count,
        "units_count": units_count,
        "tenants_count": tenants_count
    }


def generate_excel(headers, rows, sheet_name="Report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]
    
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1D68DF", end_color="1D68DF", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for row in rows:
        ws.append([str(val) if val is not None else "" for val in row])
        
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.alignment = left_align

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_pdf(headers, rows, report_title="Rental Report"):
    import html
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'),
        alignment=1,
        spaceAfter=15
    )
    
    cell_header_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=0
    )
    
    cell_body_style = ParagraphStyle(
        'TableBody',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155'),
        alignment=0
    )
    
    elements = []
    
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 10))
    
    table_data = []
    
    header_row = [Paragraph(html.escape(str(h)), cell_header_style) for h in headers]
    table_data.append(header_row)
    
    for row in rows:
        body_row = [Paragraph(html.escape(str(val)) if val is not None else "", cell_body_style) for val in row]
        table_data.append(body_row)
        
    num_cols = len(headers)
    col_width = 720.0 / num_cols
    col_widths = [col_width] * num_cols
    
    lower_title = report_title.lower()
    if "properties" in lower_title:
        col_widths = [60, 110, 140, 70, 70, 60, 65, 75, 70]
    elif "lease" in lower_title:
        col_widths = [60, 110, 65, 125, 75, 75, 70, 70, 70]
    elif "payment" in lower_title or "ledger" in lower_title:
        col_widths = [60, 110, 65, 125, 70, 70, 60, 80, 80]
    elif "maintenance" in lower_title:
        col_widths = [60, 110, 65, 125, 60, 70, 80, 70, 80]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D68DF')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ])
    t.setStyle(t_style)
    elements.append(t)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


@router.get("/reports/export")
def export_rental_report(
    type: str = Query(..., pattern="^(properties|leases|payments|maintenance)$"),
    format: str = Query("csv"),
    property_id: str = Query("all"),
    db: Session = Depends(get_rental_db),
    current_user: RentalUser = Depends(require_rental_role("super_admin", "landlord")),
):
    role_name = (current_user.role.role_name if current_user.role else "").lower()
    is_super = role_name == "super_admin"

    prop_query = db.query(Property).filter(Property.active_status == True)
    if not is_super:
        prop_query = prop_query.filter(Property.landlord_id == current_user.user_id)
    
    if property_id != "all":
        try:
            p_id = int(property_id)
            prop_query = prop_query.filter(Property.property_id == p_id)
        except ValueError:
            pass
            
    properties = prop_query.all()
    property_ids = [p.property_id for p in properties]
    
    if not property_ids:
        raise HTTPException(status_code=400, detail="No properties found to export reports.")

    headers = []
    rows = []

    # Export Properties
    if type == "properties":
        headers = ["Property ID", "Property Name", "Address", "City", "State", "Zip Code", "Total Units", "Occupied Units", "Vacant Units"]
        for p in properties:
            tot_units = db.query(Unit).filter(Unit.property_id == p.property_id, Unit.active_status == True).count()
            occ_units = db.query(Unit).filter(Unit.property_id == p.property_id, Unit.status == "OCCUPIED", Unit.active_status == True).count()
            vac_units = db.query(Unit).filter(Unit.property_id == p.property_id, Unit.status == "VACANT", Unit.active_status == True).count()
            rows.append([
                p.property_id,
                p.name,
                p.address,
                p.city or "N/A",
                p.state or "N/A",
                p.zip_code or "N/A",
                tot_units,
                occ_units,
                vac_units
            ])

    # Export Leases
    elif type == "leases":
        headers = ["Lease ID", "Property Name", "Unit Number", "Tenant Email", "Start Date", "End Date", "Rent Amount ($)", "Security Deposit ($)", "Status"]
        leases = (
            db.query(Lease)
            .filter(Lease.unit_id.in_(
                db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
            ))
            .order_by(Lease.created_date.desc())
            .all()
        )
        for l in leases:
            prop_name = l.unit.property.name if l.unit and l.unit.property else "N/A"
            unit_num = l.unit.unit_number if l.unit else "N/A"
            rows.append([
                l.lease_id,
                prop_name,
                unit_num,
                l.tenant_email or "N/A",
                l.start_date.strftime('%Y-%m-%d') if l.start_date else "N/A",
                l.end_date.strftime('%Y-%m-%d') if l.end_date else "N/A",
                l.rent_amount,
                l.security_deposit,
                l.status
            ])

    # Export Payments
    elif type == "payments":
        headers = ["Invoice ID", "Property Name", "Unit Number", "Tenant Email", "Due Date", "Amount ($)", "Status", "Payment Method", "Transaction ID"]
        ledgers = (
            db.query(RentalLedger)
            .filter(RentalLedger.lease_id.in_(
                db.query(Lease.lease_id).filter(
                    Lease.unit_id.in_(
                        db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                    )
                )
            ))
            .order_by(RentalLedger.due_date.desc())
            .all()
        )
        for led in ledgers:
            prop_name = led.lease.unit.property.name if led.lease and led.lease.unit and led.lease.unit.property else "N/A"
            unit_num = led.lease.unit.unit_number if led.lease and led.lease.unit else "N/A"
            tenant_email = led.lease.tenant_email if led.lease else "N/A"
            rows.append([
                led.invoice_id,
                prop_name,
                unit_num,
                tenant_email,
                led.due_date.strftime('%Y-%m-%d') if led.due_date else "N/A",
                led.amount,
                led.status,
                led.payment_method or "N/A",
                led.transaction_id or "N/A"
            ])

    # Export Maintenance
    elif type == "maintenance":
        headers = ["Request ID", "Property Name", "Unit Number", "Title", "Priority", "Status", "Estimated Cost ($)", "Payment Status", "Created Date"]
        maint_requests = (
            db.query(RentalMaintenanceRequest)
            .filter(RentalMaintenanceRequest.lease_id.in_(
                db.query(Lease.lease_id).filter(
                    Lease.unit_id.in_(
                        db.query(Unit.unit_id).filter(Unit.property_id.in_(property_ids), Unit.active_status == True)
                    )
                )
            ))
            .order_by(RentalMaintenanceRequest.created_date.desc())
            .all()
        )
        for req in maint_requests:
            prop_name = req.lease.unit.property.name if req.lease and req.lease.unit and req.lease.unit.property else "N/A"
            unit_num = req.lease.unit.unit_number if req.lease and req.lease.unit else "N/A"
            rows.append([
                req.request_id,
                prop_name,
                unit_num,
                req.title,
                req.priority,
                req.status,
                req.estimated_cost,
                req.payment_status,
                req.created_date.strftime('%Y-%m-%d %H:%M') if req.created_date else "N/A"
            ])

    fmt = format.lower()
    from app.services.rental.audit_service import log_rental_action
    log_rental_action(db, "GENERATE_REPORT", "rental", f"Generated and exported {fmt.upper()} report of type: '{type}'", current_user.user_id)

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    if fmt == "excel":
        sheet_title = f"{type.title()} Report"
        excel_data = generate_excel(headers, rows, sheet_name=sheet_title)
        filename = f"rental_report_{type}_{timestamp}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    elif fmt == "pdf":
        report_title = f"Rental Portfolio - {type.title()} Report"
        pdf_data = generate_pdf(headers, rows, report_title=report_title)
        filename = f"rental_report_{type}_{timestamp}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        filename = f"rental_report_{type}_{timestamp}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
