import subprocess
import sys
try:
    import openpyxl
    import reportlab
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl==3.1.2", "reportlab==4.1.0"])

import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.dependencies.auth import get_verified_user, require_role
from app.models.user import User
from app.models.violation import Violation, ViolationStatus, ViolationType
from app.models.service_request import ServiceRequest, ServiceRequestStatus, ServiceRequestType
from app.models.payment import Payment
from app.models.amenity import AmenityBooking, Amenity
from app.models.community import Community

router = APIRouter(prefix="/report", tags=["Reports"])


#  GET /api/report/{community_id}/stats
@router.get("/{community_id}/stats")
def get_report_stats(
    community_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("super_admin", "property_manager", "board_member")),
):
    role_name = current_user.role.role_name if current_user.role else None
    if role_name not in {"super_admin", "sales_admin"}:
        from app.models.user import UserCommunity
        assoc = db.query(UserCommunity).filter(
            UserCommunity.user_id == current_user.user_id,
            UserCommunity.community_id == community_id
        ).first()
        if not assoc:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to view reports for this community."
            )

    # 1. Violations Stats
    total_violations = db.query(Violation).filter(
        Violation.community_id == community_id,
        Violation.active_status == True
    ).count()

    # Counts by status
    violation_status_counts = (
        db.query(ViolationStatus.violation_status, func.count(Violation.violation_id))
        .join(Violation, Violation.violation_status_id == ViolationStatus.violation_status_id)
        .filter(Violation.community_id == community_id, Violation.active_status == True)
        .group_by(ViolationStatus.violation_status)
        .all()
    )
    status_summary = {status: count for status, count in violation_status_counts}

    # Sum of fine amount
    total_fine_amount = db.query(func.sum(Violation.amount)).filter(
        Violation.community_id == community_id,
        Violation.active_status == True
    ).scalar() or 0.0

    total_disputes = db.query(Violation).filter(
        Violation.community_id == community_id,
        Violation.is_disputed == True,
        Violation.active_status == True
    ).count()

    # 2. Service Requests Stats
    total_sr = db.query(ServiceRequest).filter(
        ServiceRequest.community_id == community_id,
        ServiceRequest.active_status == True
    ).count()

    sr_status_counts = (
        db.query(ServiceRequestStatus.status_name, func.count(ServiceRequest.request_id))
        .join(ServiceRequest, ServiceRequest.status_id == ServiceRequestStatus.status_id)
        .filter(ServiceRequest.community_id == community_id, ServiceRequest.active_status == True)
        .group_by(ServiceRequestStatus.status_name)
        .all()
    )
    sr_status_summary = {status: count for status, count in sr_status_counts}

    # 3. Payment Stats
    total_payments = db.query(Payment).filter(
        Payment.community_id == community_id,
        Payment.active_status == True
    ).count()

    total_amount_collected = db.query(func.sum(Payment.amount)).filter(
        Payment.community_id == community_id,
        Payment.status == "COMPLETED",
        Payment.active_status == True
    ).scalar() or 0.0

    payment_reasons = (
        db.query(Payment.reason, func.sum(Payment.amount))
        .filter(Payment.community_id == community_id, Payment.status == "COMPLETED", Payment.active_status == True)
        .group_by(Payment.reason)
        .all()
    )
    payment_reason_summary = {reason: amount for reason, amount in payment_reasons}

    # 4. Amenity Booking Stats
    total_bookings = db.query(AmenityBooking).filter(
        AmenityBooking.community_id == community_id,
        AmenityBooking.active_status == True
    ).count()

    amenities_count = db.query(Amenity).filter(
        Amenity.community_id == community_id,
        Amenity.active_status == True
    ).count()

    # 5. Resident Stats
    from app.models.user import Role
    residents_count = db.query(User).join(Role, User.role_id == Role.role_id).filter(
        User.community_id == community_id,
        User.active_status == True,
        ~Role.role_name.in_(["super_admin", "sales_admin", "vendor"])
    ).count()

    return {
        "violations": {
            "total": total_violations,
            "fine_amount": total_fine_amount,
            "disputed": total_disputes,
            "by_status": status_summary
        },
        "service_requests": {
            "total": total_sr,
            "by_status": sr_status_summary
        },
        "payments": {
            "total_count": total_payments,
            "total_collected": total_amount_collected,
            "by_reason": payment_reason_summary
        },
        "amenity_bookings": {
            "total": total_bookings,
            "amenities_count": amenities_count
        },
        "residents_count": residents_count
    }


def generate_excel(headers, rows, sheet_name="Report"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30] # Excel limit is 31 chars
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal-600
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Write header
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Write rows
    for row in rows:
        ws.append([str(val) if val is not None else "" for val in row])
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Align data cells and set fonts
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.alignment = left_align

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_pdf(headers, rows, report_title="Community Report"):
    import html
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    pdf_buffer = io.BytesIO()
    # Use landscape letter size (11 x 8.5 inches, or 792 x 612 pt)
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'), # slate-900
        alignment=1, # Center
        spaceAfter=15
    )
    
    cell_header_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=0 # Left
    )
    
    cell_body_style = ParagraphStyle(
        'TableBody',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155'), # slate-700
        alignment=0 # Left
    )
    
    elements = []
    
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 10))
    
    table_data = []
    
    # Wrap header cells and escape XML
    header_row = [Paragraph(html.escape(str(h)), cell_header_style) for h in headers]
    table_data.append(header_row)
    
    # Wrap body cells and escape XML
    for row in rows:
        body_row = [Paragraph(html.escape(str(val)) if val is not None else "", cell_body_style) for val in row]
        table_data.append(body_row)
        
    num_cols = len(headers)
    col_width = 720.0 / num_cols
    col_widths = [col_width] * num_cols
    
    # Adjust specific column widths to sum up to 720pt
    lower_title = report_title.lower()
    if "violation" in lower_title:
        # 11 columns: ["Violation ID", "Resident Name", "Email", "Unit No", "Violation Type", "Fine Amount ($)", "Issued Date", "Due Date", "Status", "Disputed", "Dispute Reason"]
        col_widths = [45, 75, 90, 45, 75, 55, 55, 55, 50, 45, 130]
    elif "service" in lower_title:
        # 8 columns: ["Request ID", "Resident Name", "Title", "Category", "Priority", "Status", "Created Date", "Closed Date"]
        col_widths = [50, 95, 155, 80, 50, 60, 115, 115]
    elif "payment" in lower_title:
        # 8 columns: ["Payment ID", "Payer Name", "Email", "Amount ($)", "Reason", "Payment Method", "Status", "Date"]
        col_widths = [55, 105, 130, 60, 110, 85, 75, 100]
    elif "booking" in lower_title or "amenity" in lower_title:
        # 9 columns: ["Booking ID", "Amenity Name", "Booked By", "Email", "Booking Date", "Slot", "Fee Amount ($)", "Status", "Paid"]
        col_widths = [50, 95, 85, 110, 65, 90, 75, 75, 75]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D9488')), # Teal-600
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ])
    t.setStyle(t_style)
    elements.append(t)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


#  GET /api/report/{community_id}/export
@router.get("/{community_id}/export")
def export_report(
    community_id: int,
    type: str = Query(..., pattern="^(violations|servicerequests|payments|bookings)$"),
    format: str = Query("csv"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("super_admin", "property_manager", "board_member")),
):
    role_name = current_user.role.role_name if current_user.role else None
    if role_name not in {"super_admin", "sales_admin"}:
        from app.models.user import UserCommunity
        assoc = db.query(UserCommunity).filter(
            UserCommunity.user_id == current_user.user_id,
            UserCommunity.community_id == community_id
        ).first()
        if not assoc:
            raise HTTPException(
                status_code=403,
                detail="You do not have permission to export reports for this community."
            )

    community = db.query(Community).filter(Community.community_id == community_id).first()
    community_name = community.name if community else "Community"

    headers = []
    rows = []

    # Export Violations
    if type == "violations":
        headers = ["Violation ID", "Resident Name", "Email", "Unit No", "Violation Type", "Fine Amount ($)", "Issued Date", "Due Date", "Status", "Disputed", "Dispute Reason"]
        violations = (
            db.query(Violation)
            .filter(Violation.community_id == community_id, Violation.active_status == True)
            .order_by(Violation.created_date.desc())
            .all()
        )
        for v in violations:
            resident_name = f"{v.client.first_name} {v.client.last_name}" if v.client else "N/A"
            email = v.client.email_id if v.client else "N/A"
            
            unit = "N/A"
            if v.client:
                from app.models.user import UserCommunity
                assoc = db.query(UserCommunity).filter(
                    UserCommunity.user_id == v.client.user_id,
                    UserCommunity.community_id == community_id
                ).first()
                if assoc:
                    unit = assoc.unit_no or "N/A"
                elif v.client.community_id == community_id:
                    unit = v.client.unit_no or "N/A"

            vtype = v.violation_type.name if v.violation_type else "N/A"
            rows.append([
                v.violation_id,
                resident_name,
                email,
                unit,
                vtype,
                v.amount,
                v.violation_date.strftime('%Y-%m-%d') if v.violation_date else "N/A",
                v.violation_due_date.strftime('%Y-%m-%d') if v.violation_due_date else "N/A",
                v.status.violation_status if v.status else "N/A",
                "Yes" if v.is_disputed else "No",
                v.dispute_description or ""
            ])

    # Export Service Requests
    elif type == "servicerequests":
        headers = ["Request ID", "Resident Name", "Title", "Category", "Priority", "Status", "Created Date", "Closed Date"]
        requests = (
            db.query(ServiceRequest)
            .filter(ServiceRequest.community_id == community_id, ServiceRequest.active_status == True)
            .order_by(ServiceRequest.created_date.desc())
            .all()
        )
        for r in requests:
            resident_name = f"{r.submitted_by.first_name} {r.submitted_by.last_name}" if r.submitted_by else "N/A"
            category = r.service_type.type_name if r.service_type else "N/A"
            rows.append([
                r.request_id,
                resident_name,
                r.title,
                category,
                r.priority,
                r.status.status_name if r.status else "N/A",
                r.created_date.strftime('%Y-%m-%d %H:%M') if r.created_date else "N/A",
                r.closed_date.strftime('%Y-%m-%d %H:%M') if r.closed_date else "N/A"
            ])

    # Export Payments
    elif type == "payments":
        headers = ["Payment ID", "Payer Name", "Email", "Amount ($)", "Reason", "Payment Method", "Status", "Date"]
        payments = (
            db.query(Payment)
            .filter(Payment.community_id == community_id, Payment.active_status == True)
            .order_by(Payment.payment_date.desc())
            .all()
        )
        for p in payments:
            payer_name = f"{p.user.first_name} {p.user.last_name}" if p.user else "N/A"
            email = p.user.email_id if p.user else "N/A"
            rows.append([
                p.payment_id,
                payer_name,
                email,
                p.amount,
                p.reason,
                p.payment_method or "N/A",
                p.status,
                p.payment_date.strftime('%Y-%m-%d %H:%M') if p.payment_date else "N/A"
            ])

    # Export Bookings
    elif type == "bookings":
        headers = ["Booking ID", "Amenity Name", "Booked By", "Email", "Booking Date", "Slot", "Fee Amount ($)", "Status", "Paid"]
        bookings = (
            db.query(AmenityBooking)
            .filter(AmenityBooking.community_id == community_id, AmenityBooking.active_status == True)
            .order_by(AmenityBooking.booking_date.desc())
            .all()
        )
        for b in bookings:
            amenity_name = b.amenity.name if b.amenity else "N/A"
            booked_by = f"{b.booked_by.first_name} {b.booked_by.last_name}" if b.booked_by else "N/A"
            email = b.booked_by.email_id if b.booked_by else "N/A"
            slot_desc = f"Slot 1 (8am-2pm)" if b.slot_number == 1 else f"Slot 2 (2pm-8pm)"
            rows.append([
                b.booking_id,
                amenity_name,
                booked_by,
                email,
                b.booking_date.strftime('%Y-%m-%d') if b.booking_date else "N/A",
                slot_desc,
                b.fee_amount,
                b.status,
                "Yes" if b.is_paid else "No"
            ])

    fmt = format.lower()
    
    # Log report generation activity
    from app.services.audit_service import log_action
    log_action(
        db=db,
        action="GENERATE_REPORT",
        module="community",
        description=f"Generated and exported {fmt.upper()} report of type: '{type}'",
        user_id=current_user.user_id,
        community_id=community_id,
    )

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    if fmt == "excel":
        sheet_title = f"{type.title()} Report"
        excel_data = generate_excel(headers, rows, sheet_name=sheet_title)
        filename = f"report_{type}_{community_id}_{timestamp}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    elif fmt == "pdf":
        report_title = f"{community_name} - {type.title()} Report"
        pdf_data = generate_pdf(headers, rows, report_title=report_title)
        filename = f"report_{type}_{community_id}_{timestamp}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:
        # Default to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        filename = f"report_{type}_{community_id}_{timestamp}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

